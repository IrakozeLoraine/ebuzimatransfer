"""Shared helpers for reading uploaded .csv / .xlsx spreadsheets for the bulk
import flows (units, facilities, resources).

The CSV reader auto-detects the column delimiter so files exported by Excel in
locales that use ``;`` (semicolon) — or ``\\t`` (tab) — import the same as plain
comma-separated files.
"""
from __future__ import annotations
import csv
import io
from app.core.exceptions import ValidationError


def read_xlsx_rows(file_bytes: bytes) -> list[tuple]:
    """Return the active worksheet's rows as a list of value tuples."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise ValidationError("Could not read the uploaded file. Please upload a valid .xlsx file.")
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def read_csv_rows(file_bytes: bytes) -> list[tuple]:
    """Decode and parse a CSV file into a list of row tuples.

    The delimiter (``,``, ``;`` or tab) is sniffed from the first line so that
    spreadsheets exported by Excel in semicolon locales still import correctly.
    """
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise ValidationError("Could not read the uploaded file. Please upload a valid .csv file.")

    first_line = text.split("\n", 1)[0]
    try:
        dialect = csv.Sniffer().sniff(first_line, delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        # Fall back to whichever candidate appears most in the header line.
        delimiter = max(",;\t", key=first_line.count)

    return [tuple(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]
