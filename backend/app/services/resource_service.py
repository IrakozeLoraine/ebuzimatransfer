from __future__ import annotations
import uuid
from typing import List, Optional, Sequence
from datetime import datetime
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from app.core.exceptions import NotFoundError, ResourceReservedError, ValidationError
from app.core.tiers import tier_rank
from app.models.resource import Resource, ResourceStatus, ResourceType, ResourceReservation
from app.models.unit import Unit
from app.models.facility import Facility
from app.repositories.resource_repository import ResourceRepository
from app.schemas.resource import (
    ResourceCreate,
    ResourceStatusUpdate,
    ResourceOut,
    ReservationOut,
    ResourceUsageOut,
    ResourceImportResult,
    ResourceImportError,
    CapacityRow,
    DashboardActivityRow,
)


def _to_out(resource: Resource) -> ResourceOut:
    return ResourceOut(
        id=resource.id,
        resource_name=resource.resource_name,
        resource_code=resource.resource_code,
        notes=resource.notes,
        resource_type=resource.resource_type,
        unit_id=resource.unit_id,
        facility_id=resource.facility_id,
        quantity=resource.quantity,
        status=resource.status,
        facility_name=resource.facility.name if resource.facility else None,
        unit_name=resource.unit.name if resource.unit else None,
    )


# Map of accepted resource_type values (case-insensitive) -> enum
_RESOURCE_TYPE_BY_VALUE = {t.value.lower(): t for t in ResourceType}


class ResourceService:
    def __init__(self, session: AsyncSession):
        self.repo = ResourceRepository(session)
        self.session = session

    async def _get_unit(self, unit_id: uuid.UUID) -> Unit:
        result = await self.session.execute(select(Unit).where(Unit.id == unit_id))
        unit = result.scalar_one_or_none()
        if not unit:
            raise NotFoundError("Unit")
        return unit

    async def _validate_unit_facility(
        self, unit_id: uuid.UUID | None, facility_id: uuid.UUID | None
    ) -> None:
        """A clinical unit may only be attached to a facility whose tier is at
        or above the unit's tier (cascading catalog)."""
        if unit_id is None or facility_id is None:
            return
        unit = await self._get_unit(unit_id)
        facility = await self.session.get(Facility, facility_id)
        if not facility:
            raise NotFoundError("Facility")
        if tier_rank(unit.tier) > tier_rank(facility.type):
            raise ValidationError("This unit is not available at the selected facility's tier")

    async def create(self, data: ResourceCreate) -> Resource:
        payload = data.model_dump()
        await self._validate_unit_facility(payload.get("unit_id"), payload.get("facility_id"))
        resource = Resource(**payload)
        return await self.repo.create(resource)

    async def get(self, resource_id: uuid.UUID) -> Resource:
        resource = await self.repo.get_by_id(resource_id)
        if not resource:
            raise NotFoundError("Resource")
        return resource

    async def list_scoped(
        self,
        facility_ids: Optional[Sequence[uuid.UUID]] = None,
        facility_id: Optional[uuid.UUID] = None,
        unassigned: bool = False,
        status: Optional[ResourceStatus] = None,
    ) -> List[ResourceOut]:
        resources = await self.repo.list_scoped(
            facility_ids=facility_ids, facility_id=facility_id, unassigned=unassigned, status=status
        )
        return [_to_out(r) for r in resources]

    async def list_available(self, unit_id: uuid.UUID | None = None) -> List[ResourceOut]:
        resources = await self.repo.list_available(unit_id=unit_id)
        return [_to_out(r) for r in resources]

    async def update_status(self, resource_id: uuid.UUID, data: ResourceStatusUpdate) -> Resource:
        resource = await self.get(resource_id)
        resource.status = data.status
        await self.session.flush()
        return resource

    async def assign(
        self,
        resource_id: uuid.UUID,
        facility_id: uuid.UUID | None,
        unit_id: uuid.UUID | None,
    ) -> ResourceOut:
        resource = await self.get(resource_id)
        if unit_id is not None:
            # A clinical unit only exists in the context of a facility.
            if facility_id is None:
                raise ValidationError("A facility is required when assigning a unit")
            await self._validate_unit_facility(unit_id, facility_id)
            resource.unit_id = unit_id
            resource.facility_id = facility_id
        else:
            # Assign to a facility without a specific unit, or clear (back to stock).
            resource.unit_id = None
            resource.facility_id = facility_id
        await self.session.flush()
        # Reload with relationships for display.
        result = await self.session.execute(
            select(Resource)
            .where(Resource.id == resource_id)
            .options(selectinload(Resource.unit), selectinload(Resource.facility))
        )
        return _to_out(result.scalar_one())

    async def usage(self, resource_id: uuid.UUID) -> ResourceUsageOut:
        result = await self.session.execute(
            select(Resource)
            .where(Resource.id == resource_id)
            .options(selectinload(Resource.unit), selectinload(Resource.facility))
        )
        resource = result.scalar_one_or_none()
        if not resource:
            raise NotFoundError("Resource")
        rows = await self.repo.reservations_for(resource_id)
        reservations = [
            ReservationOut(
                id=res.id,
                reserved_by=res.reserved_by,
                reserved_by_name=user.full_name,
                planned_admission_time=res.planned_admission_time,
                created_at=res.created_at,
            )
            for res, user in rows
        ]
        return ResourceUsageOut(resource=_to_out(resource), reservations=reservations)

    @staticmethod
    def _read_xlsx_rows(file_bytes: bytes) -> list[tuple]:
        import io
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception:
            raise ValidationError("Could not read the uploaded file. Please upload a valid .xlsx file.")
        ws = wb.active
        return list(ws.iter_rows(values_only=True))

    @staticmethod
    def _read_csv_rows(file_bytes: bytes) -> list[tuple]:
        import csv
        import io

        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            raise ValidationError("Could not read the uploaded file. Please upload a valid .csv file.")
        return [tuple(row) for row in csv.reader(io.StringIO(text))]

    async def import_from_excel(
        self,
        file_bytes: bytes,
        default_facility_id: uuid.UUID | None = None,
        is_csv: bool = False,
    ) -> ResourceImportResult:
        """Parse an .xlsx or .csv file and bulk-create resources.

        Expected header row (case-insensitive): resource_name, resource_code,
        resource_type, quantity, unit, notes. The ``unit`` column holds the
        clinical unit's *name*; it is resolved against the units available at
        the target facility's tier. Rows whose unit name is unknown (or not
        available at this facility) are reported as errors while the remaining
        valid rows are still imported.
        """
        rows = self._read_csv_rows(file_bytes) if is_csv else self._read_xlsx_rows(file_bytes)
        if not rows:
            return ResourceImportResult(created=0, errors=[])

        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

        def col(*names: str) -> Optional[int]:
            for name in names:
                if name in header:
                    return header.index(name)
            return None

        idx_name = col("resource_name")
        idx_code = col("resource_code")
        idx_type = col("resource_type")
        idx_qty = col("quantity")
        idx_unit = col("unit", "unit_name")
        idx_notes = col("notes")

        errors: List[ResourceImportError] = []
        if idx_name is None:
            raise ValidationError("Missing required 'resource_name' column in the spreadsheet.")

        created = 0

        # Resolve the target facility's tier so we can scope unit names to the
        # units actually available at that facility (cascading catalog). When no
        # facility is given (super-admin central stock), names resolve against the
        # full active catalog.
        facility_tier: str | None = None
        if default_facility_id is not None:
            facility = await self.session.get(Facility, default_facility_id)
            facility_tier = facility.type if facility else None

        # Build a case-insensitive name -> Unit lookup of the units available for
        # this import, computed once.
        active_units = (
            (await self.session.execute(select(Unit).where(Unit.is_active.is_(True))))
            .scalars()
            .all()
        )
        units_by_name: dict[str, Unit] = {}
        for u in active_units:
            if facility_tier is not None and tier_rank(u.tier) > tier_rank(facility_tier):
                continue  # not available at this facility's tier
            units_by_name.setdefault(u.name.strip().lower(), u)

        for i, raw in enumerate(rows[1:], start=2):  # row 1 is the header
            def cell(idx: Optional[int]) -> Optional[str]:
                if idx is None or idx >= len(raw) or raw[idx] is None:
                    return None
                return str(raw[idx]).strip()

            name = cell(idx_name)
            if not name:
                continue  # skip blank rows silently

            resource_type = None
            type_val = cell(idx_type)
            if type_val:
                resource_type = _RESOURCE_TYPE_BY_VALUE.get(type_val.lower())
                if resource_type is None:
                    errors.append(ResourceImportError(row=i, message=f"Unknown resource_type '{type_val}'"))
                    continue

            quantity = 1
            qty_val = cell(idx_qty)
            if qty_val:
                try:
                    quantity = int(float(qty_val))
                except ValueError:
                    errors.append(ResourceImportError(row=i, message=f"Invalid quantity '{qty_val}'"))
                    continue
                if quantity < 1:
                    errors.append(ResourceImportError(row=i, message="Quantity must be at least 1"))
                    continue

            facility_id = default_facility_id
            unit_id: uuid.UUID | None = None
            unit_raw = cell(idx_unit)
            if unit_raw:
                unit = units_by_name.get(unit_raw.strip().lower())
                if not unit:
                    errors.append(
                        ResourceImportError(
                            row=i,
                            message=f"Unit '{unit_raw}' is not available at this facility",
                        )
                    )
                    continue
                unit_id = unit.id
                # facility_id stays the import target (central stock when None).

            self.session.add(
                Resource(
                    resource_name=name,
                    resource_code=cell(idx_code),
                    resource_type=resource_type,
                    quantity=quantity,
                    notes=cell(idx_notes),
                    unit_id=unit_id,
                    facility_id=facility_id,
                    status=ResourceStatus.AVAILABLE,
                )
            )
            created += 1

        await self.session.flush()
        return ResourceImportResult(created=created, errors=errors)

    async def reserve(
        self,
        resource_id: uuid.UUID,
        reserved_by: uuid.UUID,
        planned_admission_time: datetime | None = None,
        referral_id: uuid.UUID | None = None,
    ) -> ResourceReservation:
        """Atomically reserve a resource using SELECT FOR UPDATE. Optionally links
        the reservation to the transfer request (referral) it fulfils."""
        resource = await self.repo.lock_for_update(resource_id)
        if not resource or resource.status != ResourceStatus.AVAILABLE:
            raise ResourceReservedError()

        resource.status = ResourceStatus.RESERVED
        reservation = ResourceReservation(
            resource_id=resource_id,
            reserved_by=reserved_by,
            planned_admission_time=planned_admission_time,
            referral_id=referral_id,
        )
        self.session.add(reservation)

        await self.session.flush()
        return reservation

    async def capacity_dashboard(
        self, facility_ids: Optional[Sequence[uuid.UUID]] = None
    ) -> List[CapacityRow]:
        rows = await self.repo.capacity_summary_raw(facility_ids=facility_ids)
        result = []
        for r in rows:
            result.append(CapacityRow(
                facility_id=r["facility_id"],
                facility=r["facility"],
                unit_type=r["unit_type"],
                total=r["total"] or 0,
                available=r["available"] or 0,
                occupied=r["occupied"] or 0,
                reserved=r["reserved"] or 0,
                out_of_service=r["out_of_service"] or 0,
            ))
        return result

    async def recent_activity(
        self, facility_ids: Optional[Sequence[uuid.UUID]] = None, limit: int = 20
    ) -> List[DashboardActivityRow]:
        rows = await self.repo.recent_reservations(facility_ids=facility_ids, limit=limit)
        result = []
        for r in rows:
            name = f"{r['first_name']} {r['last_name']}".strip() or None
            result.append(DashboardActivityRow(
                id=r["id"],
                resource_name=r["resource_name"],
                facility_name=r["facility_name"],
                unit_name=r["unit_name"],
                reserved_by_name=name,
                planned_admission_time=r["planned_admission_time"],
                created_at=r["created_at"],
            ))
        return result
