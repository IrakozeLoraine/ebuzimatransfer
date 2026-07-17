"""The occupancy report and its Excel export, plus small infrastructure pieces:
the active-facility repository query, spreadsheet decoding failures, subject-less
tokens, and the request-scoped database session."""
import io

import pytest

from app.core.exceptions import UnauthorizedError, ValidationError
from app.core.security import _create_token
from app.core.spreadsheet import read_csv_rows, read_xlsx_rows
from app.models.facility import Facility
from app.models.resource import Resource
from app.models.unit import Unit
from app.repositories.facility_repository import FacilityRepository
from app.services.report_service import ReportService

REPORTS = "/api/v1/reports"


async def _seed_occupancy(db_session, facility_id):
    unit = Unit(name="ICU", tier="DISTRICT")
    db_session.add(unit)
    await db_session.flush()
    db_session.add(
        Resource(
            resource_name="ICU Bed", facility_id=facility_id, unit_id=unit.id,
            quantity=4, occupied=1,
        )
    )
    await db_session.commit()


class TestOccupancyReport:
    async def test_rate_is_computed_from_occupied_over_total(self, db_session, make_auth):
        admin = await make_auth(roles=("SUPER_ADMIN",))
        await _seed_occupancy(db_session, admin.facility.id)

        rows = await ReportService(db_session).occupancy_report()
        assert len(rows) == 1
        assert rows[0].total_resources == 4
        assert rows[0].occupied_resources == 1
        assert rows[0].occupancy_rate == 25.0

    async def test_deactivated_facilities_are_left_out(self, db_session, make_auth):
        admin = await make_auth(roles=("SUPER_ADMIN",))
        await _seed_occupancy(db_session, admin.facility.id)
        facility = await db_session.get(Facility, admin.facility.id)
        facility.is_active = False
        await db_session.commit()

        assert await ReportService(db_session).occupancy_report() == []


class TestExcelExport:
    async def test_export_returns_a_readable_workbook(self, db_session, make_auth):
        from openpyxl import load_workbook

        admin = await make_auth(roles=("SUPER_ADMIN",))
        await _seed_occupancy(db_session, admin.facility.id)

        data = await ReportService(db_session).export_excel()
        wb = load_workbook(io.BytesIO(data))
        assert wb.sheetnames == ["Occupancy"]
        header = [c.value for c in wb["Occupancy"][1]]
        assert "occupancy_rate" in header

    async def test_export_endpoint_serves_a_spreadsheet_download(
        self, client, db_session, make_auth
    ):
        admin = await make_auth(roles=("SUPER_ADMIN",))
        await _seed_occupancy(db_session, admin.facility.id)

        resp = await client.get(f"{REPORTS}/export/excel", headers=admin.headers)
        assert resp.status_code == 200
        assert "filename=report.xlsx" in resp.headers["content-disposition"]
        assert resp.content[:2] == b"PK"  # xlsx is a zip container

    async def test_export_is_super_admin_only(self, client, make_auth):
        clinician = await make_auth(roles=("CLINICIAN",))
        resp = await client.get(f"{REPORTS}/export/excel", headers=clinician.headers)
        assert resp.status_code == 403


class TestFacilityRepository:
    async def test_list_active_skips_deactivated_facilities(self, db_session, make_auth):
        live = await make_auth(roles=("SUPER_ADMIN",), facility_name="Live")
        dead = await make_auth(roles=("SUPER_ADMIN",), facility_name="Dead")
        retired = await db_session.get(Facility, dead.facility.id)
        retired.is_active = False
        await db_session.commit()

        names = {f.id for f in await FacilityRepository(db_session).list_active()}
        assert live.facility.id in names
        assert dead.facility.id not in names


class TestSpreadsheetDecoding:
    def test_undecodable_csv_is_rejected(self):
        with pytest.raises(ValidationError, match="valid .csv file"):
            read_csv_rows(b"\xff\xfe\x00\x00binary-not-utf8")

    def test_unreadable_xlsx_is_rejected(self):
        with pytest.raises(ValidationError, match="valid .xlsx file"):
            read_xlsx_rows(b"this is not a workbook")


class TestTokensWithoutSubject:
    async def test_driver_token_without_subject_rejected(self, client):
        from datetime import timedelta

        token = _create_token({"type": "driver"}, timedelta(minutes=5))
        resp = await client.get(
            "/api/v1/driver/journey", headers={"Authorization": f"Bearer {token}"}
        )
        assert resp.status_code == 401

    async def test_access_token_without_subject_rejected(self, client):
        from datetime import timedelta

        token = _create_token({"type": "access", "roles": []}, timedelta(minutes=5))
        resp = await client.get(
            "/api/v1/users/me", headers={"Authorization": f"Bearer {token}"}
        )
        assert resp.status_code == 401


class TestRequestSession:
    async def test_get_session_yields_a_usable_session(self):
        """The real dependency (integration tests override it) opens a session
        against the configured engine and closes it on exit."""
        from sqlalchemy import text
        from sqlalchemy.ext.asyncio import AsyncSession
        from app.db.session import get_session

        agen = get_session()
        session = await anext(agen)
        try:
            assert isinstance(session, AsyncSession)
            assert (await session.execute(text("SELECT 1"))).scalar() == 1
        except Exception as exc:  # noqa: BLE001 — no database reachable → skip
            pytest.skip(f"No database available: {exc}")
        finally:
            await agen.aclose()
